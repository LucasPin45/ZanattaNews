"""
telegram_news_bot.py

Monitor Parlamentar Zanatta - Versão Completa
==============================================
FONTES:
- RSS de portais de notícias
- Google Alerts (via RSS)
- Pauta da Câmara dos Deputados (API Dados Abertos)
- Diário Oficial da União (scraping)

FUNCIONALIDADES:
- Coleta paralela de múltiplas fontes
- Priorização por relevância
- Deduplicação inteligente
- Configuração externa via YAML
- Logging estruturado
- Retry automático
- Resumo diário
"""

import os
import sys
import re
import time
import sqlite3
import html
import logging
import argparse
from datetime import datetime, timezone, timedelta
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import quote, urljoin
from typing import Optional

import yaml
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import feedparser
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup


# ============================================================
# LOGGING
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger(__name__)


# ============================================================
# CONFIGURAÇÃO
# ============================================================
def load_config(config_path: str = "config.yaml") -> dict:
    """Carrega configuração do arquivo YAML."""
    path = Path(config_path)
    if not path.exists():
        logger.error(f"Arquivo de configuração não encontrado: {config_path}")
        raise SystemExit(1)
    
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ============================================================
# TELEGRAM
# ============================================================
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
    raise SystemExit("Defina TELEGRAM_BOT_TOKEN e TELEGRAM_CHAT_ID nas variáveis de ambiente.")

BRT = timezone(timedelta(hours=-3))


# ============================================================
# HTTP SESSION COM RETRY
# ============================================================
def create_session() -> requests.Session:
    """Cria sessão HTTP com retry automático."""
    session = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=0.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"]
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })
    return session


SESSION = create_session()


# ============================================================
# BANCO DE DADOS
# ============================================================
def init_db(db_path: str):
    """Inicializa banco SQLite com índices."""
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sent (
            id TEXT PRIMARY KEY,
            title TEXT,
            link TEXT,
            source TEXT,
            source_type TEXT,
            published_at TEXT,
            sent_at TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sent_published ON sent(published_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sent_sent_at ON sent(sent_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sent_source_type ON sent(source_type)")
    conn.commit()
    conn.close()


def was_sent(db_path: str, item_id: str) -> bool:
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM sent WHERE id=?", (item_id,))
    r = cur.fetchone()
    conn.close()
    return r is not None


def mark_sent(db_path: str, item: dict):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO sent VALUES (?, ?, ?, ?, ?, ?, ?)",
        (
            item["id"],
            item["title"],
            item["link"],
            item["source"],
            item.get("source_type", "rss"),
            item["published_at"],
            datetime.now(timezone.utc).isoformat(),
        )
    )
    conn.commit()
    conn.close()


def get_sent_titles_today(db_path: str) -> list[str]:
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0).isoformat()
    cur.execute("SELECT title FROM sent WHERE sent_at >= ?", (today_start,))
    titles = [row[0] for row in cur.fetchall()]
    conn.close()
    return titles


def count_sent_today(db_path: str) -> dict:
    """Conta notícias enviadas hoje por tipo."""
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    today = datetime.now(BRT).strftime("%Y-%m-%d")
    cur.execute("""
        SELECT source_type, COUNT(*) 
        FROM sent 
        WHERE DATE(sent_at) = ? 
        GROUP BY source_type
    """, (today,))
    counts = dict(cur.fetchall())
    conn.close()
    return counts


# ============================================================
# HISTÓRICO XLSX
# ============================================================
def init_xlsx(xlsx_path: str):
    if os.path.exists(xlsx_path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Historico"
    ws.append([
        "Enviado_BRT", "Publicado_BRT", "Fonte", "Tipo",
        "Titulo", "Link", "Keyword", "Contexto", "Relevancia"
    ])
    wb.save(xlsx_path)


def append_xlsx(xlsx_path: str, row: list):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.append(row)
    wb.save(xlsx_path)


# ============================================================
# TELEGRAM
# ============================================================
def send_telegram(msg: str) -> bool:
    """Envia mensagem para o Telegram."""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    try:
        r = SESSION.post(
            url,
            json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": msg,
                "parse_mode": "HTML",
                "disable_web_page_preview": True,
            },
            timeout=25,
        )
        r.raise_for_status()
        return True
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro ao enviar Telegram: {e}")
        return False


def send_daily_summary(db_path: str):
    """Envia resumo diário."""
    counts = count_sent_today(db_path)
    total = sum(counts.values())
    
    breakdown = []
    type_emoji = {"rss": "📰", "google_alert": "🔔", "camara": "🏛️", "dou": "📜"}
    type_name = {"rss": "Notícias", "google_alert": "Google Alerts", "camara": "Câmara", "dou": "DOU"}
    
    for t, c in counts.items():
        emoji = type_emoji.get(t, "📌")
        name = type_name.get(t, t)
        breakdown.append(f"{emoji} {name}: {c}")
    
    msg = (
        f"📊 <b>Resumo do dia</b>\n\n"
        f"📰 Total enviado: <b>{total}</b>\n\n"
        + "\n".join(breakdown) + "\n\n"
        f"🕐 {datetime.now(BRT).strftime('%d/%m/%Y %H:%M')} (BRT)"
    )
    
    if send_telegram(msg):
        logger.info(f"Resumo diário enviado: {total} itens")


# ============================================================
# AUXILIARES
# ============================================================
def normalize(s: str) -> str:
    return (s or "").strip()


def parse_published_dt(entry) -> Optional[datetime]:
    if hasattr(entry, "published_parsed") and entry.published_parsed:
        return datetime(*entry.published_parsed[:6], tzinfo=timezone.utc)
    if hasattr(entry, "updated_parsed") and entry.updated_parsed:
        return datetime(*entry.updated_parsed[:6], tzinfo=timezone.utc)
    return None


def find_context(title: str, summary: str, keywords: list[str], context_chars: int, max_len: int) -> tuple[str, str]:
    """Encontra contexto ao redor da keyword."""
    # Primeiro tenta no summary
    raw = normalize(summary)
    low = raw.lower()

    for kw in keywords:
        k = kw.lower()
        idx = low.find(k)
        if idx != -1:
            start = max(0, idx - context_chars)
            end = min(len(raw), idx + len(k) + context_chars)
            ctx = raw[start:end]
            
            ctx_low = ctx.lower()
            kw_idx = ctx_low.find(k)
            if kw_idx != -1:
                ctx = ctx[:kw_idx] + f"<b>{ctx[kw_idx:kw_idx+len(k)]}</b>" + ctx[kw_idx+len(k):]
            
            if len(ctx) > max_len:
                ctx = ctx[:max_len] + "…"
            return kw, html.escape(ctx, quote=False).replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
    
    # Se não achou no summary, tenta no título
    title_low = title.lower()
    for kw in keywords:
        if kw.lower() in title_low:
            return kw, "(ver título)"
    
    return "", ""


def count_keywords(text: str, keywords: list[str]) -> int:
    text_lower = text.lower()
    return sum(1 for k in keywords if k.lower() in text_lower)


def is_similar(title1: str, title2: str, threshold: float = 0.85) -> bool:
    return SequenceMatcher(None, title1.lower(), title2.lower()).ratio() > threshold


def is_duplicate_by_similarity(title: str, existing_titles: list[str], threshold: float) -> bool:
    for existing in existing_titles:
        if is_similar(title, existing, threshold):
            return True
    return False


# ============================================================
# FETCH RSS (PARALELO)
# ============================================================
def fetch_feed(url: str) -> tuple[str, list, str]:
    try:
        parsed = feedparser.parse(url)
        source = parsed.feed.title if hasattr(parsed.feed, "title") else url
        return (url, parsed.entries[:120], source)
    except Exception as e:
        logger.warning(f"Erro no feed {url}: {e}")
        return (url, [], None)


def fetch_all_feeds(feeds: list[str], workers: int) -> list[tuple[str, list, str]]:
    results = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(fetch_feed, url): url for url in feeds}
        for future in as_completed(futures):
            try:
                result = future.result()
                if result[1]:
                    results.append(result)
            except Exception as e:
                logger.warning(f"Exceção em feed: {e}")
    return results


# ============================================================
# CÂMARA DOS DEPUTADOS - PAUTA DO PLENÁRIO
# ============================================================
def fetch_camara_pauta() -> list[dict]:
    """Busca pauta do dia da Câmara via API."""
    items = []
    today = datetime.now(BRT).strftime("%Y-%m-%d")
    
    try:
        # API de eventos
        url = f"https://dadosabertos.camara.leg.br/api/v2/eventos?dataInicio={today}&dataFim={today}&ordem=ASC&ordenarPor=dataHoraInicio"
        r = SESSION.get(url, timeout=30)
        r.raise_for_status()
        data = r.json()
        
        for evento in data.get("dados", []):
            # Filtra apenas plenário
            orgaos = evento.get("orgaos", [])
            is_plenario = any("Plenário" in o.get("nome", "") for o in orgaos)
            
            if not is_plenario:
                continue
            
            titulo = evento.get("descricaoTipo", "Sessão")
            descricao = evento.get("descricao", "")
            data_hora = evento.get("dataHoraInicio", "")
            situacao = evento.get("descricaoSituacao", "")
            
            # Busca pauta do evento
            evento_id = evento.get("id")
            pauta_items = []
            
            if evento_id:
                try:
                    pauta_url = f"https://dadosabertos.camara.leg.br/api/v2/eventos/{evento_id}/pauta"
                    pr = SESSION.get(pauta_url, timeout=15)
                    if pr.status_code == 200:
                        pauta_data = pr.json()
                        for p in pauta_data.get("dados", [])[:5]:  # Top 5 itens
                            prop = p.get("proposicao_", {})
                            if prop:
                                pauta_items.append(f"• {prop.get('siglaTipo', '')} {prop.get('numero', '')}/{prop.get('ano', '')}")
                except:
                    pass
            
            # Formata horário
            hora_fmt = ""
            if data_hora:
                try:
                    dt = datetime.fromisoformat(data_hora.replace("Z", "+00:00"))
                    hora_fmt = dt.astimezone(BRT).strftime("%H:%M")
                except:
                    hora_fmt = data_hora
            
            items.append({
                "id": f"camara_{evento_id}",
                "title": f"{titulo}: {descricao}" if descricao else titulo,
                "link": f"https://www.camara.leg.br/evento-legislativo/{evento_id}" if evento_id else "https://www.camara.leg.br/agenda",
                "source": "Câmara dos Deputados",
                "source_type": "camara",
                "published_at": data_hora,
                "hora": hora_fmt,
                "situacao": situacao,
                "pauta": pauta_items,
            })
        
        logger.info(f"Câmara: {len(items)} eventos encontrados")
        
    except Exception as e:
        logger.error(f"Erro ao buscar pauta da Câmara: {e}")
    
    return items


def send_camara_pauta(db_path: str):
    """Envia pauta do dia da Câmara."""
    eventos = fetch_camara_pauta()
    
    if not eventos:
        logger.info("Câmara: nenhum evento no plenário hoje")
        return
    
    # Monta mensagem consolidada
    msg_parts = ["🏛️ <b>PAUTA DO PLENÁRIO - CÂMARA</b>\n"]
    msg_parts.append(f"📅 {datetime.now(BRT).strftime('%d/%m/%Y')}\n")
    
    for ev in eventos:
        msg_parts.append(f"\n⏰ <b>{ev['hora']}</b> - {html.escape(ev['title'])}")
        if ev['situacao']:
            msg_parts.append(f"   📌 {ev['situacao']}")
        if ev['pauta']:
            msg_parts.append("   📋 Pauta:")
            for p in ev['pauta']:
                msg_parts.append(f"      {p}")
        msg_parts.append(f"   🔗 {ev['link']}")
    
    msg = "\n".join(msg_parts)
    
    # Verifica se já enviou hoje
    pauta_id = f"camara_pauta_{datetime.now(BRT).strftime('%Y%m%d')}"
    if was_sent(db_path, pauta_id):
        logger.info("Pauta da Câmara já enviada hoje")
        return
    
    if send_telegram(msg):
        mark_sent(db_path, {
            "id": pauta_id,
            "title": "Pauta do Plenário",
            "link": "https://www.camara.leg.br/agenda",
            "source": "Câmara dos Deputados",
            "source_type": "camara",
            "published_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info("Pauta da Câmara enviada com sucesso")


# ============================================================
# DIÁRIO OFICIAL DA UNIÃO (DOU)
# ============================================================
def fetch_dou(keywords: list[str], secoes: list[str]) -> list[dict]:
    """Busca publicações no DOU via site da Imprensa Nacional."""
    items = []
    today = datetime.now(BRT).strftime("%d-%m-%Y")
    
    for keyword in keywords:
        for secao in secoes:
            try:
                # URL de busca do DOU
                search_url = "https://www.in.gov.br/consulta/-/buscar/dou"
                params = {
                    "q": keyword,
                    "s": secao,
                    "exactDate": "personalizado",
                    "publishFrom": today,
                    "publishTo": today,
                    "delta": 20,
                    "sortType": "0"
                }
                
                r = SESSION.get(search_url, params=params, timeout=30)
                if r.status_code != 200:
                    continue
                
                soup = BeautifulSoup(r.text, "html.parser")
                
                # Busca resultados
                results = soup.select(".resultados-dou .resultado-item") or soup.select(".results-list .item")
                
                for item in results[:10]:  # Max 10 por keyword
                    try:
                        # Tenta extrair título e link
                        title_el = item.select_one("h5, .title, a.title-content")
                        link_el = item.select_one("a[href*='/web/dou/']")
                        
                        if not title_el:
                            continue
                        
                        title = title_el.get_text(strip=True)
                        link = ""
                        if link_el:
                            href = link_el.get("href", "")
                            link = urljoin("https://www.in.gov.br", href)
                        
                        # Extrai resumo se disponível
                        summary_el = item.select_one(".abstract, .resumo, p")
                        summary = summary_el.get_text(strip=True) if summary_el else ""
                        
                        # Extrai órgão
                        org_el = item.select_one(".orgao, .organization")
                        orgao = org_el.get_text(strip=True) if org_el else "DOU"
                        
                        item_id = f"dou_{secao}_{hash(title + link)}"
                        
                        items.append({
                            "id": item_id,
                            "title": title,
                            "link": link or f"https://www.in.gov.br/consulta/-/buscar/dou?q={quote(keyword)}",
                            "summary": summary,
                            "source": f"DOU Seção {secao} - {orgao}",
                            "source_type": "dou",
                            "published_at": datetime.now(timezone.utc).isoformat(),
                            "published_brt": datetime.now(BRT).strftime("%d/%m %H:%M"),
                            "kw": keyword,
                        })
                    except Exception as e:
                        logger.debug(f"Erro ao processar item DOU: {e}")
                        continue
                
            except Exception as e:
                logger.warning(f"Erro ao buscar DOU para '{keyword}': {e}")
                continue
    
    # Remove duplicatas
    seen = set()
    unique_items = []
    for item in items:
        if item["title"] not in seen:
            seen.add(item["title"])
            unique_items.append(item)
    
    logger.info(f"DOU: {len(unique_items)} publicações encontradas")
    return unique_items


# ============================================================
# PROCESSAMENTO DE RSS E GOOGLE ALERTS
# ============================================================
def process_rss_items(feed_results: list, keywords: list, blocklist: list, 
                      cutoff: datetime, context_chars: int, max_context_len: int,
                      source_type: str = "rss") -> list[dict]:
    """Processa itens de feeds RSS."""
    items = []
    
    for url, entries, source in feed_results:
        for e in entries:
            title = normalize(getattr(e, "title", ""))
            link = normalize(getattr(e, "link", ""))
            summary = normalize(getattr(e, "summary", ""))

            if not title or not link:
                continue

            blob = f"{title}\n{summary}".lower()
            
            if not any(k.lower() in blob for k in keywords):
                continue
            
            if any(b.lower() in blob for b in blocklist):
                continue

            pub_dt = parse_published_dt(e)
            if pub_dt and pub_dt < cutoff:
                continue

            kw, ctx = find_context(title, summary, keywords, context_chars, max_context_len)
            relevance = count_keywords(blob, keywords)

            items.append({
                "id": getattr(e, "id", link),
                "title": title,
                "link": link,
                "source": source,
                "source_type": source_type,
                "published_at": pub_dt.isoformat() if pub_dt else None,
                "published_brt": pub_dt.astimezone(BRT).strftime("%d/%m %H:%M") if pub_dt else "",
                "kw": kw,
                "ctx": ctx,
                "relevance": relevance,
            })
    
    return items


# ============================================================
# EXECUÇÃO PRINCIPAL
# ============================================================
def run(config: dict, hours_override: int = None, send_summary: bool = False, 
        pauta_only: bool = False):
    """Executa o bot."""
    
    settings = config["settings"]
    files = config["files"]
    
    db_path = files["database"]
    xlsx_path = files["history_xlsx"]
    
    init_db(db_path)
    init_xlsx(xlsx_path)
    
    # Modo resumo diário
    if send_summary:
        send_daily_summary(db_path)
        return
    
    # Modo pauta da Câmara
    if pauta_only:
        if config.get("camara_pauta", {}).get("enabled", False):
            send_camara_pauta(db_path)
        return
    
    lookback_hours = hours_override or settings["lookback_hours"]
    max_items = settings["max_items_per_run"]
    sleep_time = settings["sleep_between_sends"]
    context_chars = settings["context_chars"]
    max_context_len = settings["max_context_len"]
    similarity_threshold = settings["similarity_threshold"]
    parallel_workers = settings["parallel_workers"]
    
    keywords = config["keywords"]
    blocklist = config["blocklist"]
    cutoff = datetime.now(timezone.utc) - timedelta(hours=lookback_hours)
    
    all_items = []
    
    # 1. RSS de notícias
    feeds = config.get("feeds", [])
    if feeds:
        logger.info(f"Buscando RSS ({len(feeds)} feeds)...")
        feed_results = fetch_all_feeds(feeds, parallel_workers)
        rss_items = process_rss_items(feed_results, keywords, blocklist, cutoff, 
                                       context_chars, max_context_len, "rss")
        all_items.extend(rss_items)
        logger.info(f"RSS: {len(rss_items)} itens")
    
    # 2. Google Alerts
    ga_config = config.get("google_alerts", {})
    if ga_config.get("enabled", False) and ga_config.get("feeds"):
        logger.info("Buscando Google Alerts...")
        ga_feeds = ga_config["feeds"]
        ga_results = fetch_all_feeds(ga_feeds, parallel_workers)
        ga_items = process_rss_items(ga_results, keywords, blocklist, cutoff,
                                      context_chars, max_context_len, "google_alert")
        all_items.extend(ga_items)
        logger.info(f"Google Alerts: {len(ga_items)} itens")
    
    # 3. Diário Oficial da União
    dou_config = config.get("dou", {})
    if dou_config.get("enabled", False):
        logger.info("Buscando DOU...")
        dou_keywords = dou_config.get("keywords", [])
        dou_secoes = dou_config.get("secoes", ["1"])
        dou_items = fetch_dou(dou_keywords, dou_secoes)
        
        # Adiciona contexto e relevância
        for item in dou_items:
            item["ctx"] = item.get("summary", "")[:max_context_len]
            item["relevance"] = 1
            if not item.get("published_brt"):
                item["published_brt"] = datetime.now(BRT).strftime("%d/%m %H:%M")
        
        all_items.extend(dou_items)
        logger.info(f"DOU: {len(dou_items)} itens")
    
    # 4. Pauta da Câmara (verifica se é horário)
    camara_config = config.get("camara_pauta", {})
    if camara_config.get("enabled", False):
        hora_atual = datetime.now(BRT).hour
        hora_pauta = camara_config.get("send_hour", 7)
        if hora_atual == hora_pauta:
            send_camara_pauta(db_path)
    
    logger.info(f"Total de itens: {len(all_items)}")
    
    # Ordena por relevância e data
    all_items.sort(key=lambda x: (x.get("relevance", 0), x.get("published_at") or ""), reverse=True)
    
    # Envia itens
    sent_titles = get_sent_titles_today(db_path)
    sent = 0
    skipped_similar = 0
    skipped_duplicate = 0
    
    for it in all_items:
        if sent >= max_items:
            break
        
        if was_sent(db_path, it["id"]):
            skipped_duplicate += 1
            continue
        
        if is_duplicate_by_similarity(it["title"], sent_titles, similarity_threshold):
            skipped_similar += 1
            continue
        
        # Emoji por tipo de fonte
        type_emoji = {
            "rss": "📰",
            "google_alert": "🔔",
            "dou": "📜",
            "camara": "🏛️"
        }
        emoji = type_emoji.get(it.get("source_type", "rss"), "📰")
        
        msg = (
            f"{emoji} <b>{html.escape(it['title'])}</b>\n"
            f"🏷 <i>{html.escape(it['source'])} • {it.get('published_brt', '')} (BRT)</i>\n"
            f"🔎 <b>Gatilho:</b> <code>{html.escape(it.get('kw', ''))}</code>\n"
            f"🧾 <i>{it.get('ctx', '')}</i>\n"
            f"🔗 {it['link']}"
        )

        if send_telegram(msg):
            append_xlsx(xlsx_path, [
                datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S"),
                it.get("published_brt", ""),
                it["source"],
                it.get("source_type", "rss"),
                it["title"],
                it["link"],
                it.get("kw", ""),
                it.get("ctx", ""),
                it.get("relevance", 0),
            ])

            mark_sent(db_path, it)
            sent_titles.append(it["title"])
            sent += 1
            time.sleep(sleep_time)

    logger.info(
        f"Concluído: {sent} enviadas, "
        f"{skipped_duplicate} duplicadas, "
        f"{skipped_similar} similares"
    )


def main():
    parser = argparse.ArgumentParser(description="Telegram News Bot - Monitor Parlamentar")
    parser.add_argument("--hours", type=int, help="Override lookback hours")
    parser.add_argument("--config", default="config.yaml", help="Path to config file")
    parser.add_argument("--summary", action="store_true", help="Send daily summary only")
    parser.add_argument("--pauta", action="store_true", help="Send Câmara pauta only")
    args = parser.parse_args()
    
    config = load_config(args.config)
    run(config, hours_override=args.hours, send_summary=args.summary, pauta_only=args.pauta)


if __name__ == "__main__":
    main()